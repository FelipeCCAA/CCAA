"""
Carga el maestro de productos desde `Recetas_Cod_Producto.xlsx`.

Por qué un comando y no una migración de datos: los productos son datos de
**negocio**, no catálogo del sistema. Una migración los sembraría también en
la base de pruebas —igual que pasa con los documentos de liberación y los
equipos— y ahí no pintan nada: ninguna prueba habla de la Leche Descremada MH
de Soprole, y tenerla presente solo abre la puerta a que una prueba pase por
un producto que no creó.

**El SKU se compone, no se copia.** La columna «SKU» del archivo se ignora a
propósito: trae filas mal codificadas (§4.2 de `SKU_PRODUCTOS.md`) y copiarla
metería en el maestro exactamente el defecto que `Producto.save()` existe para
evitar — un código que contradice los atributos del mismo producto.

Por omisión **simula**: imprime lo que haría y no escribe nada. Se aplica con
`--aplicar`. Es carga de maestro sobre una base viva; ver antes qué va a pasar
vale más que el paso extra.

    python manage.py cargar_productos
    python manage.py cargar_productos --aplicar
"""

from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


from maestros.models import Mandante, Producto


ARCHIVO = (
    Path(__file__).resolve().parents[4]
    / "docs"
    / "levantamiento-2026-07"
    / "Recetas_Cod_Producto.xlsx"
)

HOJA = "Código de Producto"

#: Primera fila con datos y columnas (0-indexadas dentro de la fila) del bloque
#: de productos, que empieza en la columna P. El archivo mezcla los catálogos
#: (columnas A-N) con la tabla de productos en la misma hoja.
PRIMERA_FILA = 3
COL_NATURALEZA, COL_CLIENTE, COL_CATEGORIA, COL_TIPO, COL_FORMATO = 15, 16, 17, 18, 19
COL_NOMBRE = 28

NATURALEZA = {
    "servicio terceros": "servicio_terceros",
    "producto propio": "producto_propio",
}

CLIENTE = {
    "nestlé": "nestle",
    "nestle": "nestle",
    "colun": "colun",
    "soprole": "soprole",
    "cliente no definido": "no_definido",
}

CATEGORIA = {
    "leche en polvo": "leche_polvo",
    "pre condensados": "precondensado",
    "crema": "crema",
    "matequilla": "mantequilla",          # así está escrito en el archivo
    "mantequilla": "mantequilla",
    "materiales diversos": "materiales_diversos",
    "suero": "suero",
    "extracto de malta": "extracto_malta",
    "leche en polvo instantanea": "lp_instantanea",
    "leche en polvo c/lec": "lp_con_lecitina",
    "leche fluida": "leche_fluida",
}

TIPO = {
    "entera": "entera",
    "semidescremada": "semidescremada",
    "descremada": "descremada",
    "con sal": "con_sal",
    "sin sal": "sin_sal",
    "no definido": "no_definido",
    "estandarizada": "estandarizada",
}

FORMATO = {
    "granel": "granel",
    "saco 25kg": "saco_25kg",
    "caja 20kg": "caja_20kg",
}

#: A qué familia pertenece cada categoría. La familia decide qué documentos del
#: Dossier le aplican (`DocumentoLiberacion.aplica_a`), así que no es cosmética.
FAMILIA = {
    "leche_polvo": "polvo",
    "lp_instantanea": "polvo",
    "lp_con_lecitina": "polvo",
    "suero": "polvo",
    "extracto_malta": "polvo",
    "precondensado": "liquido",
    "leche_fluida": "liquido",
    "crema": "crema",
    "mantequilla": "otro",
    "materiales_diversos": "otro",
}

#: Correcciones al archivo, por nombre de producto. Están aquí y no aplicadas a
#: mano para que se vea qué se cambió y por qué: el archivo es la fuente, y
#: apartarse de él sin dejar rastro es como llegó a tener estos errores.
CORRECCIONES = {
    "Leche Entera Estándar 28% NE 25kg": {
        "categoria": "leche_polvo",
        "motivo": (
            "El archivo la codifica con Categoría = «Crema» (§4.2). Es una "
            "leche en polvo; el atributo está mal, no solo el SKU."
        ),
    },
    "Leche Entera c/LdS 27% SP 25kg": {
        "categoria": "lp_con_lecitina",
        "motivo": (
            "Lleva lecitina de soya y la categoría 11 ya existe. El archivo la "
            "deja como 02 y así colisiona con la leche entera sin lecitina."
        ),
    },
    "Leche Descremada c/LdS MH SP 25kg": {
        "categoria": "lp_con_lecitina",
        "motivo": "Mismo caso: c/LdS con categoría 02 en el archivo.",
    },
}


class _Simulacion(Exception):
    """Corta la transacción de la vista previa. No es un error."""


def _clave(mapa, valor, campo, fila):
    texto = str(valor or "").strip().lower()
    clave = mapa.get(texto)

    if clave is None:
        raise CommandError(
            f"Fila {fila}: no sé traducir {campo} = {valor!r}. Agrega la "
            f"equivalencia en cargar_productos.py — inventarla metería un "
            f"producto mal clasificado en el maestro."
        )

    return clave


class Command(BaseCommand):
    help = "Carga el maestro de productos desde el Excel del levantamiento."

    def add_arguments(self, parser):
        parser.add_argument(
            "--aplicar",
            action="store_true",
            help="Escribe en la base. Sin esto solo simula.",
        )

    def handle(self, *args, **opciones):
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise CommandError(
                "Falta openpyxl. Instálalo con: pip install openpyxl"
            ) from error

        if not ARCHIVO.exists():
            raise CommandError(f"No encuentro el archivo: {ARCHIVO}")

        filas = self._leer(load_workbook)
        aplicar = opciones["aplicar"]
        resultados = []

        # La simulación recorre **el mismo camino** y revierte al final, en vez
        # de calcular aparte lo que pasaría. Una segunda implementación del
        # cálculo es libre de discrepar justo con la que escribe, y entonces la
        # vista previa deja de servir para lo único que sirve.
        try:
            with transaction.atomic():
                resultados = [self._cargar(f) for f in filas]

                if not aplicar:
                    raise _Simulacion
        except _Simulacion:
            pass

        self._informar(resultados, aplicar)

    # ------------------------------------------------------------ lectura

    def _leer(self, load_workbook):
        libro = load_workbook(ARCHIVO, data_only=True)
        hoja = libro[HOJA]
        filas = []

        for numero, fila in enumerate(
            hoja.iter_rows(min_row=PRIMERA_FILA, values_only=True), PRIMERA_FILA
        ):
            if not fila[COL_NATURALEZA]:
                continue

            nombre = str(fila[COL_NOMBRE] or "").strip()

            if not nombre:
                raise CommandError(f"Fila {numero}: producto sin nombre.")

            datos = {
                "fila": numero,
                "nombre": nombre,
                "naturaleza_comercial": _clave(
                    NATURALEZA, fila[COL_NATURALEZA], "naturaleza", numero
                ),
                "cliente": _clave(CLIENTE, fila[COL_CLIENTE], "cliente", numero),
                "categoria": _clave(
                    CATEGORIA, fila[COL_CATEGORIA], "categoría", numero
                ),
                "tipo": _clave(TIPO, fila[COL_TIPO], "tipo", numero),
                "formato": _clave(FORMATO, fila[COL_FORMATO], "formato", numero),
                "corregido": None,
            }

            correccion = CORRECCIONES.get(nombre)

            if correccion:
                datos["corregido"] = (
                    f"categoría {datos['categoria']} → {correccion['categoria']}: "
                    f"{correccion['motivo']}"
                )
                datos["categoria"] = correccion["categoria"]

            filas.append(datos)

        return filas

    # ------------------------------------------------------------ escritura

    def _mandante(self, clave):
        """
        El mandante que corresponde a ese código de cliente.

        Devuelve uno solo sin desempatar nada: la restricción
        `mandante_unico_por_codigo_cliente` garantiza que no haya dos. Hubo un
        tiempo en que sí los había —«Nestle» y «Nestlé» a la vez— y este
        método elegía el más antiguo y avisaba; ese aviso ya no puede
        dispararse, así que no está.
        """
        mandante = Mandante.objects.filter(codigo_cliente=clave).first()

        if mandante is not None:
            return mandante

        nombre = {"nestle": "Nestlé", "colun": "Colun", "soprole": "Soprole"}.get(
            clave, "CCAA"
        )

        return Mandante.objects.create(nombre=nombre, codigo_cliente=clave)

    def _atributos(self, datos, mandante):
        return {
            "mandante": mandante,
            "naturaleza_comercial": datos["naturaleza_comercial"],
            "categoria": datos["categoria"],
            "tipo": datos["tipo"],
            "formato": datos["formato"],
            "mercado": Producto.Mercado.LOCAL,
            "familia": FAMILIA[datos["categoria"]],
            # `naturaleza` queda en el valor por omisión —terminado— para todos.
            # Dónde está cada uno en la cadena lo dice su receta, y no hay
            # ninguna cargada: deducirlo aquí sería inventar la cadena de la
            # planta. Mientras tanto, un producto sin receta se reporta como
            # cadena incompleta, que es la señal correcta.
        }

    def _cargar(self, datos):
        mandante = self._mandante(datos["cliente"])

        producto, creado = Producto.objects.update_or_create(
            nombre=datos["nombre"],
            mandante=mandante,
            defaults=self._atributos(datos, mandante),
        )

        return {**datos, "sku": producto.codigo, "creado": creado}

    # ------------------------------------------------------------ informe

    def _informar(self, resultados, aplicar):
        titulo = "APLICADO" if aplicar else "SIMULACIÓN (usa --aplicar para escribir)"
        self.stdout.write(self.style.MIGRATE_HEADING(f"\n{titulo}\n"))

        for r in resultados:
            marca = "+" if r["creado"] else "="
            self.stdout.write(f"  {marca} {r['sku']:<16} {r['nombre']}")

        corregidos = [r for r in resultados if r["corregido"]]

        if corregidos:
            self.stdout.write(
                self.style.WARNING("\nCorregidos respecto del archivo:")
            )
            for r in corregidos:
                self.stdout.write(f"  · {r['nombre']}\n      {r['corregido']}")

        self._informar_colisiones(resultados)

    def _informar_colisiones(self, resultados):
        """
        Dos productos con el mismo SKU es lo que decide si hace falta el 7.º
        segmento (`SKU_PRODUCTOS.md` §4.1). Se cuenta después de aplicar las
        correcciones, que es justamente lo que había que medir.
        """
        por_sku = defaultdict(list)

        for r in resultados:
            por_sku[r["sku"]].append(r["nombre"])

        colisiones = {s: n for s, n in por_sku.items() if len(n) > 1}

        if not colisiones:
            self.stdout.write(self.style.SUCCESS("\nSin colisiones de SKU."))
            return

        self.stdout.write(
            self.style.ERROR(
                f"\n{len(colisiones)} SKU compartidos por más de un producto. "
                "Son los que necesitan que negocio decida: o son el mismo "
                "producto repetido en el archivo, o hace falta el 7.º segmento "
                "de variante."
            )
        )

        for sku, nombres in sorted(colisiones.items()):
            self.stdout.write(f"  {sku}")
            for nombre in nombres:
                self.stdout.write(f"      · {nombre}")
